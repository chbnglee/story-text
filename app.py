"""
Story Tool — Streamlit Web App
  Step 1. 음원 생성   : Excel/CSV 업로드 → Normal / Easy / Difficult / mBook xlsx 생성
  Step 2. ICMS txt 생성: xlsx 3종 업로드 → 감정 태그 포함 TXT 생성 (Gemini API 사용)
"""

import io
import json
import re
import zipfile
from collections import defaultdict

import streamlit as st
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ─────────────────────────────────────────────────────────────────────────────
# 공통 상수 & 스타일
# ─────────────────────────────────────────────────────────────────────────────

LEVELS = ["Normal", "Easy", "Difficult"]

HEADER_FILL = PatternFill("solid", fgColor="4F81BD")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
CELL_FONT   = Font(name="Calibri", size=11)
WRAP_ALIGN  = Alignment(wrap_text=True, vertical="top")
THIN_SIDE   = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(
    left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE
)
COL_WIDTHS = {"ID": 12, "Key": 22, "Text": 80}

SCENE_RE = re.compile(r"#SC(\d+)", re.MULTILINE)

EMOTION_PROMPT = """\
You are tagging the emotional tone of each scene in a children's story.

Available tags: Neutral, Happy, Sad, Angry
Rules:
- Every scene value MUST start with "Neutral".
- You MUST assign at least one non-Neutral tag (Happy, Sad, or Angry) to AT LEAST 3 scenes.
- Look for: crying / loss / disappointment → Sad | cheering / success / reunion → Happy | frustration / conflict → Angry
- Multiple tags are allowed (e.g. "Neutral, Sad, Happy" for bittersweet moments).
- Err toward tagging — a subtle emotional cue is enough.

Return ONLY a JSON object. Keys = scene IDs, values = comma-separated tag strings.
Example:
{
  "SC01": "Neutral",
  "SC02": "Neutral, Sad",
  "SC03": "Neutral",
  "SC04": "Neutral, Happy",
  "SC05": "Neutral, Angry, Sad"
}

Story scenes (Normal version):
"""

CHUNK_SYSTEM_PROMPT = """\
You are an expert in English lexical chunking for language education.
Split English sentences into lexical chunks by placing "@@" at chunk boundaries.
Base all decisions on Thornbury (2019) "Learning Language in Chunks" (Cambridge Papers in ELT).

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
THEORETICAL FOUNDATION  (Thornbury 2019, p.5)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Chunks:
  (1) consist of more than one word
  (2) are conventionalised — co-occur more than by chance
  (3) exhibit VARYING DEGREES of fixedness
  (4) exhibit VARYING DEGREES of idiomaticity
  (5) are likely processed as single units

CRITICAL: Properties (3) and (4) are matters of DEGREE, not binary checks.
Do NOT use a fixed threshold like "must satisfy N properties."
Use CATEGORY MEMBERSHIP (C1–C9 below) as the primary decision criterion.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CHUNK CATEGORIES  (Thornbury 2019, p.3, p.5)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
C1  COLLOCATION
    Two or more words that frequently co-occur.
    Examples: "heavy rain", "take time", "burst into tears",
              "draw a picture", "a little afraid", "with delight"
    Test: Do these specific words combine with markedly higher frequency
          than alternatives? (Verifiable via COCA, SkELL; MI ≥ 3.0)

C2  FIXED EXPRESSION
    Invariant multi-word units, often discourse markers.
    Examples: "by the way", "in fact", "as a result",
              "once upon a time", "in the end", "every day"
    Test: Does altering any word make the form unacceptable? If YES → fixed.

C3  FORMULAIC UTTERANCE
    Whole utterances retrieved as single units in recognisable situations.
    Examples: "no way!", "never mind", "good luck", "take your time",
              "long time no see", "way to go"
    Test: Is the entire utterance conventionally produced as one unit?

C4  SENTENCE STARTER
    Patterns with open substitutable slots that begin utterances.
    Examples: "Have you ever ___?", "What about ___?", "It's time to ___"
    Test: Does the pattern recur with different content filling the slot?

C5  VERB PATTERN
    Verb-centred patterns with fixed peripheral elements and open slots.
    Examples: "make/fight one's way", "take ___ for granted"
    Test: Does the verb appear in a recurring structural frame?

C6  IDIOM
    Non-compositional figurative expressions.
    Examples: "a wild goose chase", "run out of steam",
              "once in a blue moon", "plain sailing"
    Test: Is meaning unpredictable from individual word meanings alone?

C7  PHRASAL VERB
    Verb + particle (adverb or preposition) with a unified meaning.
    Examples: "take off", "look for", "look at", "give up",
              "come across", "search for", "pick up", "run out of"
    Test: Does the verb+particle carry a unified meaning often
          non-derivable from the parts independently?

C8  FUNCTIONAL EXPRESSION
    Formulaic exponents of recognisable speech acts.
    Examples: "Would you like ___?" (offering),
              "I'm sorry to hear that" (sympathy)
    Test: Does this perform a recognisable speech act in conventional form?

C9  CLUSTER / BUNDLE
    High-frequency n-grams that recur as fixed sequences in corpora.
    Examples: "at the end of the", "every day", "you know what",
              "I was going to", "there is no way of knowing"
    Test: Per Biber et al. (1999), does this recur as a fixed sequence
          in spoken/written corpora regardless of idiomaticity?

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
DECISION PROCEDURE  (apply in order for every sentence)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Step 1 — CANDIDATE SCAN
  Identify all multi-word spans that might match categories C1–C9.

Step 2 — CATEGORY MATCH TEST
  For each candidate span, ask: "Does this clearly belong to one of C1–C9?"
    Clear match   → CHUNK. Proceed to Step 3.
    Borderline    → apply Step 2b.
    No match      → NOT a chunk. Move on.

Step 2b — BORDERLINE DEGREE ASSESSMENT
  For weak or borderline candidates, assess degree on three dimensions:
    Q1. Fixedness:          Does the form resist substitution or reordering?
    Q2. Idiomaticity:       Is meaning non-compositional?
    Q3. Conventionalisation: Does this exact combination recur in standard use?

  HIGH on ≥ 1 dimension AND moderate on another → treat as CHUNK.
  LOW on all dimensions                          → FREE COMBINATION (not a chunk).

  NOTE: Productive grammatical patterns — "be + Adj", "V + to + V",
  "the + Adj + N", "modal + V", "be + V-ing/V-ed" — score LOW on all
  dimensions even when frequent. FREQUENCY ALONE IS NOT SUFFICIENT.

Step 3 — BOUNDARY PLACEMENT
  • @@ BEFORE a chunk when the chunk follows other material.
  • @@ AFTER  a chunk when the chunk precedes other material.
  • C7 phrasal verb + object:
      @@ goes AFTER the phrasal verb and BEFORE its object.
        "Liam looks for@@Bear"   ← NOT "Liam@@looks for Bear"
  • Quoted speech + reporting clause:
      @@ between the quoted string and the reporting clause (both directions).
        "\"Bear!\"@@he calls."     (quote BEFORE reporting clause)
        "It says@@\"Bear.\""       (verb BEFORE quote)

Step 4 — JUSTIFICATION CHECK  (internal, run before writing any @@)
  Ask: "Which chunk does this @@ delimit? Which category C1–C9 does it belong to?"
  If you cannot name both the chunk and its category → REMOVE the @@.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CRITICAL EXCLUSIONS — NEVER PLACE @@ HERE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
E1  Subject–Verb boundary
    WRONG: "Liam@@draws a picture"  /  "He@@hugs Bear"
    Reason: Syntactic parsing, not lexical chunking.

E2  Auxiliary + main verb  (progressive, passive)
    WRONG: "is@@wearing a collar"  /  "is@@printed on it"
    Reason: Tense/aspect/voice = grammatical structure, not a chunk.

E3  Modal + verb
    WRONG: "could@@be a monster"  /  "may@@come tomorrow"
    Reason: Modal constructions = grammatical structure.

E4  V + to + V  (infinitive complement)
    WRONG: "begins@@to fall"  /  "wants@@to leave"
    Reason: Infinitive complementation = grammatical structure.

E5  Article + adjective + noun  (free productive noun phrase)
    WRONG: "a@@big dog"  /  "the@@blue collar"
    Reason: Free combination — not conventionalised as a fixed unit.

E6  Inside a confirmed chunk
    WRONG: "feels@@a little@@afraid"  (if "a little afraid" is treated as one chunk)
    RIGHT:  "feels@@a little afraid"
    Reason: @@ marks chunk EDGES, not internal positions.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PREPOSITIONAL PHRASE POLICY — Policy A  (Conservative)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Place @@ at a PP boundary ONLY when the PP qualifies as:
  (a) A Fixed Expression or Cluster (C2/C9):
        "in fact", "at home", "in the end", "every day", "once upon a time"
  (b) The object of a confirmed Phrasal Verb (C7):
        "looks for@@Bear"  /  "searches for@@the key"
  (c) A high-frequency manner/time collocation (C1, MI ≥ 3.0):
        "with delight", "with excitement", "in silence"

Do NOT place @@ before generic locative, temporal, or recipient PPs:
  "under a bench", "from the sky", "for Bear", "to Mom", "at the door",
  "in the forest", "from the shadows" — even after a complete clause.

Apply Policy A UNIFORMLY across all sentences in the batch.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CONSISTENCY VERIFICATION  (run before producing final output)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. Group sentences by structural pattern:
     S+V  /  S+V+O  /  S+V+Adj  /  S+V+PP  /  reporting clauses  /  quoted speech
2. Verify identical @@ treatment within each group.
3. If inconsistent → apply the more conservative (fewer splits) treatment to ALL.
4. Specifically verify:
   • No subject-verb splits anywhere
   • All reporting clauses treated identically
   • All instances of the same phrasal verb split identically
   • All time expressions of the same type split identically
   • All PPs of the same type follow Policy A

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
EXAMPLES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
── C2/C9 time expression ──
Input:  SC01_ST01_N = They play every day.
Output: SC01_ST01_N_C = They play@@every day.
  → "every day" = C2 Fixed Expression / C9 Cluster. No S/V split (E1).

── C7 phrasal verb + object ──
Input:  SC02_ST01_N = Liam looks for Bear.
Output: SC02_ST01_N_C = Liam looks for@@Bear.
  → "looks for" = C7 Phrasal Verb. @@ after PV, before object.

── No chunks → no split ──
Input:  SC03_ST01_N = Liam smiles gently at him.
Output: SC03_ST01_N_C = Liam smiles gently at him.
  → No span matches C1–C9 (Step 4: cannot name chunk + category). No @@ placed.

── E2: auxiliary construction, never split ──
Input:  SC04_ST01_D = The dog is wearing a blue collar.
Output: SC04_ST01_D_C = The dog is wearing a blue collar.
  → "is wearing" = E2 (be + V-ing). Not a chunk. "a blue collar" = free NP (E5).

── E4 + E1 + Policy A: no splits ──
Input:  SC05_ST01_D = Heavy rain begins to fall from the sky.
Output: SC05_ST01_D_C = Heavy rain begins to fall from the sky.
  → "begins to fall" = E4 (V + to + V). Not a chunk.
  → "from the sky" = generic locative PP (Policy A: no split).
  → "Heavy rain" = C1 collocation, but E1 blocks @@ at the S/V boundary.

── C3 whole-sentence formulaic utterance → no split ──
Input:  SC06_ST01_E = Take your time.
Output: SC06_ST01_E_C = Take your time.
  → Whole-sentence C3 Formulaic Utterance → return as-is.

── C2 + C9 + C1 ──
Input:  SC07_ST01_N = By the way, there is no way of knowing what will take place.
Output: SC07_ST01_N_C = By the way,@@there is no way of knowing@@what will@@take place.
  → "By the way" = C2 Fixed Expression.
  → "there is no way of knowing" = C9 Cluster / C3 Formulaic Utterance.
  → "take place" = C1 Collocation.

── C1 degree collocation + Policy A ──
Input:  SC08_ST01_D = The bear feels a little afraid in the forest.
Output: SC08_ST01_D_C = The bear feels@@a little afraid in the forest.
  → "a little afraid" = C1 Collocation (conventionalised degree + adj unit).
  → "in the forest" = generic locative PP (Policy A: no split).

── Reporting clause: quote BEFORE clause (Step 3) ──
Input:  SC09_ST01_N = "Bear!" he calls from the shadows.
Output: SC09_ST01_N_C = "Bear!"@@he calls from the shadows.
  → @@ between quoted speech and reporting clause (Step 3).
  → "from the shadows" = generic locative PP (Policy A: no split).

── Reporting clause: verb BEFORE quote (Step 3) ──
Input:  SC10_ST01_N = It says "Bear."
Output: SC10_ST01_N_C = It says@@"Bear."
  → Verb before quote → @@ between verb and quoted string (Step 3).

── E2/E4 auxiliary chain + C7 + C2/C9 ──
Input:  SC11_ST01_D = The man seems to be searching for something every day.
Output: SC11_ST01_D_C = The man seems to be searching for@@something@@every day.
  → "seems to be searching" = E2 + E4 auxiliary chain. Not a chunk.
  → "searching for" = C7 Phrasal Verb. @@ after PV, before object.
  → "every day" = C2/C9. @@ before it.

── S + be + Adj: productive, not a chunk ──
Input:  SC12_ST01_N = Lily was happy and the sun was bright.
Output: SC12_ST01_N_C = Lily was happy and the sun was bright.
  → "was happy", "was bright" = productive S + be + Adj.
  → Low on all degree dimensions (Step 2b). No @@ placed.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
OUTPUT RULES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
For every input line "KEY = sentence", output EXACTLY ONE line: "KEY_C = chunked sentence".
• Keep ALL original words — nothing deleted or added.
• Punctuation stays attached to the word immediately before it.
• No spaces before or after @@.
• If no chunks are identified, return the sentence exactly as-is.
Output ONLY the result lines — NO explanations, NO blank lines, NO extra text.
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
INPUT:
"""

# ─────────────────────────────────────────────────────────────────────────────
# 공통 유틸
# ─────────────────────────────────────────────────────────────────────────────

def safe_fname(s: str) -> str:
    """파일명에 쓸 수 없는 문자 제거."""
    return re.sub(r'[\\/*?:"<>|]', "_", s)


def _style_ws(ws):
    """헤더·데이터 행 스타일 일괄 적용."""
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font      = CELL_FONT
            cell.border    = THIN_BORDER
            cell.alignment = WRAP_ALIGN
    for idx, col in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(idx)
        ].width = COL_WIDTHS[col]
    ws.row_dimensions[1].height = 18


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 로직
# ─────────────────────────────────────────────────────────────────────────────

def split_sentences(text: str) -> list[str]:
    raw = re.split(r'(?<=[.!?"])\s+', text.strip())
    return [s.strip() for s in raw if s.strip()]


def parse_scenes(raw: str) -> list[tuple[int, list[str]]]:
    parts = SCENE_RE.split(raw.strip())
    result, i = [], 1
    while i < len(parts) - 1:
        sc_num     = int(parts[i])
        sc_text    = parts[i + 1].strip()
        sentences  = split_sentences(sc_text)
        if sentences:
            result.append((sc_num, sentences))
        i += 2
    return result


def make_level_xlsx(story_id: str, level_text: str, initial: str) -> bytes:
    """Normal / Easy / Difficult 중 하나의 xlsx 바이트 반환."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = initial
    ws.append(["ID", "Key", "Text"])
    for sc_num, sentences in parse_scenes(level_text):
        for si, sent in enumerate(sentences, start=1):
            ws.append([story_id, f"SC{sc_num:02d}_ST{si:02d}_{initial}", sent])
    _style_ws(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_mbook_xlsx(story_id: str, mbook_text: str) -> bytes:
    """mBook 텍스트를 문장 단위로 분할하여 xlsx로 저장 (Normal/Easy/Difficult와 동일 구조)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "mBook"
    ws.append(["ID", "Key", "Text"])

    scenes = parse_scenes(mbook_text)
    if scenes:
        # #SC 마커가 있으면 씬·문장 구조로 분할
        for sc_num, sentences in scenes:
            for si, sent in enumerate(sentences, start=1):
                ws.append([story_id, f"SC{sc_num:02d}_ST{si:02d}_M", sent])
    else:
        # 마커 없으면 전체를 문장 단위로 순번 부여
        for si, sent in enumerate(split_sentences(mbook_text), start=1):
            ws.append([story_id, f"ST{si:02d}_M", sent])

    _style_ws(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def read_input_file(file_bytes: bytes, filename: str) -> list[dict]:
    """xlsx 또는 csv 업로드를 읽어 dict 목록 반환."""
    if filename.lower().endswith(".csv"):
        import csv, io as _io
        for enc in ("utf-8-sig", "utf-8", "cp949", "latin-1"):
            try:
                text = file_bytes.decode(enc)
                reader = csv.DictReader(_io.StringIO(text))
                return [dict(r) for r in reader]
            except (UnicodeDecodeError, LookupError):
                continue
        raise ValueError("CSV 인코딩을 자동으로 감지할 수 없습니다.")
    else:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        headers = [
            str(c.value).strip() if c.value is not None else f"col{i}"
            for i, c in enumerate(ws[1])
        ]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            rows.append({
                headers[i]: (str(row[i]) if row[i] is not None else "")
                for i in range(len(headers))
            })
        return rows


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 로직
# ─────────────────────────────────────────────────────────────────────────────

def read_xlsx_rows(file_bytes: bytes) -> list[tuple[str, str, str]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append((str(row[0]), str(row[1]), str(row[2])))
    return rows


def extract_scene_texts(rows: list[tuple[str, str, str]]) -> dict[str, str]:
    scenes: dict[str, list[str]] = defaultdict(list)
    for _, key, text in rows:
        scenes[key.split("_")[0]].append(text)
    return {sc: " ".join(sents) for sc, sents in scenes.items()}


def _gemini_call(client, model: str, prompt: str) -> str:
    return client.models.generate_content(model=model, contents=prompt).text.strip()


def get_emotions(client, model: str, scene_texts: dict[str, str]) -> dict[str, str]:
    block = "\n\n".join(f"{k}:\n{v}" for k, v in sorted(scene_texts.items()))
    raw   = _gemini_call(client, model, EMOTION_PROMPT + block)

    def parse(r: str) -> dict:
        m = re.search(r"\{[\s\S]*\}", r)
        if not m:
            raise ValueError("JSON을 찾을 수 없습니다.")
        return json.loads(m.group())

    data = parse(raw)
    non_neutral = [k for k, v in data.items() if v.strip() != "Neutral"]
    if len(non_neutral) < 3:
        retry = (
            f"Your previous answer tagged only {len(non_neutral)} scene(s) as non-Neutral. "
            f"The requirement is AT LEAST 3 scenes with Happy, Sad, or Angry.\n"
            f"Go through every scene again and find emotional cues. "
            f"Return the complete JSON with at least 3 non-Neutral tags.\n\n"
            f"Story scenes:\n{block}"
        )
        data = parse(_gemini_call(client, model, retry))
    return data


def build_txt_content(
    level_rows: dict[str, list[tuple[str, str, str]]],
    emotions: dict[str, str],
    scene_keys: list[str],
) -> str:
    lines = []
    for level in LEVELS:
        for _, key, text in level_rows.get(level, []):
            lines.append(f"{key} = {text}")
        lines.append("")
    for sc in sorted(scene_keys):
        lines.append(f"{sc}_Emotion = {emotions.get(sc, 'Neutral')}")
    return "\n".join(lines)


def group_xlsx_files(
    files: list,
) -> dict[str, dict[str, bytes]]:
    """업로드된 파일들을 기본 이름(base)과 레벨로 그룹화."""
    pat = re.compile(r"^(.+?)_(Normal|Easy|Difficult)\.xlsx$", re.IGNORECASE)
    groups: dict[str, dict[str, bytes]] = defaultdict(dict)
    unmatched = []
    for f in files:
        m = pat.match(f.name)
        if m:
            base  = m.group(1)
            level = m.group(2).capitalize()
            groups[base][level] = f.read()
        else:
            unmatched.append(f.name)
    return dict(groups), unmatched


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 로직
# ─────────────────────────────────────────────────────────────────────────────

def parse_txt_for_chunking(content: str) -> list[tuple[str, str]]:
    """TXT 파일에서 SC##_ST##_N/E/D 키만 추출 (감정 라인 제외)."""
    pairs: list[tuple[str, str]] = []
    key_re = re.compile(r"^(SC\d+_ST\d+_[NED])\s*=\s*(.+)$")
    for line in content.splitlines():
        m = key_re.match(line.strip())
        if m:
            pairs.append((m.group(1), m.group(2).strip()))
    return pairs


def chunk_sentences_with_gemini(
    client, model: str, pairs: list[tuple[str, str]]
) -> dict[str, str]:
    """Gemini에 전체 문장 목록을 보내 청크 분할 결과를 받아온다."""
    if not pairs:
        return {}

    input_block = "\n".join(f"{k} = {v}" for k, v in pairs)
    response    = _gemini_call(client, model, CHUNK_SYSTEM_PROMPT + input_block)

    result: dict[str, str] = {}
    chunk_re = re.compile(r"^(SC\d+_ST\d+_[NED])_C\s*=\s*(.+)$")
    for line in response.splitlines():
        m = chunk_re.match(line.strip())
        if m:
            result[m.group(1)] = m.group(2).strip()
    return result


def build_chunked_txt(pairs: list[tuple[str, str]], chunked: dict[str, str]) -> str:
    """청크 결과로 출력 TXT 구성 (레벨별 빈 줄 구분)."""
    lines: list[str] = []
    prev_level: str | None = None
    for key, original in pairs:
        level = key.split("_")[2]          # N, E, or D
        if prev_level is not None and level != prev_level:
            lines.append("")
        prev_level = level
        text = chunked.get(key, original)  # fallback: 원문 그대로
        lines.append(f"{key}_C = {text}")
    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — Vocab 로직
# ─────────────────────────────────────────────────────────────────────────────

_IRREGULAR_VERBS: dict[str, set[str]] = {
    "be":         {"is","are","am","was","were","been","being"},
    "have":       {"has","had","having"},
    "do":         {"does","did","done","doing"},
    "go":         {"goes","went","gone","going"},
    "get":        {"gets","got","gotten","getting"},
    "make":       {"makes","made","making"},
    "take":       {"takes","took","taken","taking"},
    "see":        {"sees","saw","seen","seeing"},
    "come":       {"comes","came","coming"},
    "know":       {"knows","knew","known","knowing"},
    "think":      {"thinks","thought","thinking"},
    "give":       {"gives","gave","given","giving"},
    "find":       {"finds","found","finding"},
    "tell":       {"tells","told","telling"},
    "say":        {"says","said","saying"},
    "leave":      {"leaves","left","leaving"},
    "feel":       {"feels","felt","feeling"},
    "bring":      {"brings","brought","bringing"},
    "keep":       {"keeps","kept","keeping"},
    "begin":      {"begins","began","begun","beginning"},
    "hold":       {"holds","held","holding"},
    "write":      {"writes","wrote","written","writing"},
    "stand":      {"stands","stood","standing"},
    "hear":       {"hears","heard","hearing"},
    "spend":      {"spends","spent","spending"},
    "meet":       {"meets","met","meeting"},
    "run":        {"runs","ran","running"},
    "fall":       {"falls","fell","fallen","falling"},
    "sit":        {"sits","sat","sitting"},
    "win":        {"wins","won","winning"},
    "catch":      {"catches","caught","catching"},
    "buy":        {"buys","bought","buying"},
    "lose":       {"loses","lost","losing"},
    "send":       {"sends","sent","sending"},
    "build":      {"builds","built","building"},
    "show":       {"shows","showed","shown","showing"},
    "fly":        {"flies","flew","flown","flying"},
    "cry":        {"cries","cried","crying"},
    "swim":       {"swims","swam","swum","swimming"},
    "sing":       {"sings","sang","sung","singing"},
    "eat":        {"eats","ate","eaten","eating"},
    "drink":      {"drinks","drank","drunk","drinking"},
    "grow":       {"grows","grew","grown","growing"},
    "throw":      {"throws","threw","thrown","throwing"},
    "forget":     {"forgets","forgot","forgotten","forgetting"},
    "break":      {"breaks","broke","broken","breaking"},
    "choose":     {"chooses","chose","chosen","choosing"},
    "hide":       {"hides","hid","hidden","hiding"},
    "wake":       {"wakes","woke","woken","waking"},
    "wear":       {"wears","wore","worn","wearing"},
    "draw":       {"draws","drew","drawn","drawing"},
    "sleep":      {"sleeps","slept","sleeping"},
    "pay":        {"pays","paid","paying"},
    "understand": {"understands","understood","understanding"},
    "put":        {"puts","putting"},
    "cut":        {"cuts","cutting"},
    "set":        {"sets","setting"},
    "let":        {"lets","letting"},
    "hit":        {"hits","hitting"},
    "read":       {"reads","reading"},
    "lead":       {"leads","led","leading"},
    "teach":      {"teaches","taught","teaching"},
    "mean":       {"means","meant","meaning"},
    "light":      {"lights","lit","lighting"},
    "fight":      {"fights","fought","fighting"},
    "bring":      {"brings","brought","bringing"},
    "think":      {"thinks","thought","thinking"},
    "buy":        {"buys","bought","buying"},
}


def _get_all_forms(word: str) -> set[str]:
    """Return base word + all regular/irregular inflected forms."""
    w = word.strip().lower()
    forms: set[str] = {w}

    if w in _IRREGULAR_VERBS:
        forms |= _IRREGULAR_VERBS[w]
        return forms

    # Plural / 3rd-person singular
    if w.endswith(("s", "x", "z", "ch", "sh")):
        forms.add(w + "es")
    elif w.endswith("y") and len(w) > 1 and w[-2] not in "aeiou":
        forms.add(w[:-1] + "ies")
    elif not w.endswith("s"):
        forms.add(w + "s")

    # Past tense / past participle
    if w.endswith("e"):
        forms.add(w + "d")
    elif w.endswith("y") and len(w) > 1 and w[-2] not in "aeiou":
        forms.add(w[:-1] + "ied")
    elif (len(w) >= 3
          and w[-1] not in "aeiouwxy"
          and w[-2] in "aeiou"
          and w[-3] not in "aeiou"):
        forms.add(w + w[-1] + "ed")   # consonant doubling
    else:
        forms.add(w + "ed")

    # Present participle (-ing)
    if w.endswith("ie"):
        forms.add(w[:-2] + "ying")
    elif w.endswith("e") and not w.endswith("ee"):
        forms.add(w[:-1] + "ing")
    elif (len(w) >= 3
          and w[-1] not in "aeiouwxy"
          and w[-2] in "aeiou"
          and w[-3] not in "aeiou"):
        forms.add(w + w[-1] + "ing")  # consonant doubling
    else:
        forms.add(w + "ing")

    return forms


def mark_vocab(text: str, words: list[str]) -> str:
    """Wrap each target word (and all inflected forms) with ## markers."""
    if not words:
        return text
    all_forms: set[str] = set()
    for w in words:
        w = w.strip()
        if w:
            all_forms |= _get_all_forms(w)
    if not all_forms:
        return text
    sorted_forms = sorted(all_forms, key=len, reverse=True)
    pattern = r"\b(?:" + "|".join(re.escape(f) for f in sorted_forms) + r")\b"
    return re.sub(pattern, lambda m: f"##{m.group()}##", text, flags=re.IGNORECASE)


def build_vocab_txt(
    pairs: list[tuple[str, str]],
    vocab_n: list[str],
    vocab_e: list[str],
    vocab_d: list[str],
) -> str:
    """Apply ## vocab markers per level and build output TXT."""
    vocab_map = {"N": vocab_n, "E": vocab_e, "D": vocab_d}
    lines: list[str] = []
    prev_level: str | None = None
    for key, text in pairs:
        level = key.split("_")[2]          # N, E, or D
        if prev_level is not None and level != prev_level:
            lines.append("")
        prev_level = level
        marked = mark_vocab(text, vocab_map.get(level, []))
        lines.append(f"{key} = {marked}")
    return "\n".join(lines)


def parse_emotion_lines(content: str) -> list[str]:
    """TXT에서 SC##_Emotion = ... 라인만 추출."""
    emotion_re = re.compile(r"^SC\d+_Emotion\s*=\s*.+$")
    return [
        line.strip()
        for line in content.splitlines()
        if emotion_re.match(line.strip())
    ]


# ─────────────────────────────────────────────────────────────────────────────
# 페이지 설정
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Story Tool",
    page_icon="📖",
    layout="centered",
)

st.markdown("""
<style>
/* 탭 폰트 강조 */
.stTabs [data-baseweb="tab"] { font-size: 1.05rem; font-weight: 600; padding: 8px 24px; }
/* 버튼 폰트 통일 (기본보다 1pt 작게) */
div.stButton > button, div.stDownloadButton > button {
    font-size: 0.8rem !important;
}
</style>
""", unsafe_allow_html=True)

# 세션 상태 초기화
for key in ("step1_zip", "step2_result", "step3_chunk_result", "step3_vocab_result"):
    if key not in st.session_state:
        st.session_state[key] = None

tab1, tab2, tab3 = st.tabs([
    "Step 1. 음원 추출용 xlsx",
    "Step 2. 기준 txt (Source)",
    "Step 3. Chunk & Vocab txt (ICMS)",
])


# ═════════════════════════════════════════════════════════
# STEP 1 — 음원 생성
# ═════════════════════════════════════════════════════════
with tab1:
    st.subheader("Step 1. 음원 추출용 xlsx")
    st.caption(
        "**ID / Title / Normal / Easy / Difficult / mBook** 열이 있는 "
        "Excel(.xlsx) 또는 CSV 파일을 업로드하세요."
    )

    st.markdown("""
<div style="background:#e8f4fd; border-left:4px solid #4F81BD;
            padding:10px 14px; border-radius:4px; font-size:0.82rem; line-height:1.7;">
📌 Normal / Easy / Difficult 열: #SC01 기준으로 문장 단위로 분할됩니다.<br>
📌 mBook 열: #SC01 마커가 있으면 씬·문장 구조로, 없으면 문장 순번(ST01_M …)으로 분할됩니다.
</div>
""", unsafe_allow_html=True)

    st.write("")

    _template_csv = "ID,Title,Normal,Easy,Difficult,mBook\n"
    col_tmpl, _ = st.columns([1, 1])
    with col_tmpl:
        st.download_button(
            label="📋 입력 포맷 CSV 다운로드",
            data=_template_csv.encode("utf-8-sig"),
            file_name="story_input_format.csv",
            mime="text/csv",
            use_container_width=True,
            key="step1_template",
        )

    st.write("")

    uploaded_input = st.file_uploader(
        "파일 선택 (.xlsx / .csv)",
        type=["xlsx", "csv"],
        key="step1_file",
    )

    st.write("")

    col_run, col_dl = st.columns([1, 1])

    with col_run:
        run_btn = st.button("📂 음원 추출용 xlsx 생성", use_container_width=True, key="step1_run")

    if run_btn and uploaded_input:
        with st.spinner("처리 중..."):
            try:
                raw_bytes = uploaded_input.read()
                rows = read_input_file(raw_bytes, uploaded_input.name)

                zip_buf = io.BytesIO()
                counts  = {"Normal": 0, "Easy": 0, "Difficult": 0, "mBook": 0}

                with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                    for row in rows:
                        sid   = row.get("ID", "").strip()
                        title = row.get("Title", "").strip()
                        if not sid or not title:
                            continue
                        st_title = safe_fname(title)

                        for col_name, initial in [
                            ("Normal", "N"), ("Easy", "E"), ("Difficult", "D")
                        ]:
                            raw = row.get(col_name, "").strip()
                            if raw:
                                data = make_level_xlsx(sid, raw, initial)
                                zf.writestr(f"{sid}_{st_title}_{col_name}.xlsx", data)
                                counts[col_name] += 1

                        mbook = row.get("mBook", "").strip()
                        if mbook:
                            data = make_mbook_xlsx(sid, mbook)
                            zf.writestr(f"{sid}_{st_title}_mBook.xlsx", data)
                            counts["mBook"] += 1

                zip_buf.seek(0)
                st.session_state.step1_zip = zip_buf.getvalue()

                total = sum(counts.values())
                st.success(
                    f"✅ 완료! {len(rows)}개 스토리 → "
                    f"Normal {counts['Normal']}개 / Easy {counts['Easy']}개 / "
                    f"Difficult {counts['Difficult']}개 / mBook {counts['mBook']}개 "
                    f"(총 {total}개 파일)"
                )

            except Exception as e:
                st.error(f"오류가 발생했습니다: {e}")
                st.session_state.step1_zip = None

    if st.session_state.step1_zip:
        with col_dl:
            st.download_button(
                label="⬇️ ZIP 다운로드",
                data=st.session_state.step1_zip,
                file_name="script_output.zip",
                mime="application/zip",
                use_container_width=True,
                key="step1_dl",
            )


# ═════════════════════════════════════════════════════════
# STEP 2 — ICMS용 txt 생성
# ═════════════════════════════════════════════════════════
with tab2:
    st.subheader("Step 2. 기준 txt (Source)")
    st.caption(
        "**ID_Title_Normal.xlsx / ID_Title_Easy.xlsx / ID_Title_Difficult.xlsx** "
        "파일을 업로드하면 감정 태그가 포함된 TXT 파일을 생성합니다. "
        "여러 스토리를 동시에 처리할 수 있습니다."
    )

    # ── API 키 ──────────────────────────────────────────
    _secret_key = ""
    try:
        _secret_key = st.secrets.get("GEMINI_API_KEY", "")
    except Exception:
        pass

    if _secret_key:
        api_key = _secret_key
        st.success("✅ Gemini API 키가 설정되어 있습니다.", icon="🔑")
    else:
        api_key = st.text_input(
            "🔑 Gemini API Key",
            type="password",
            placeholder="AIza...",
            help="https://aistudio.google.com/app/apikey 에서 무료 발급",
        )

    # ── 파일 업로드 ──────────────────────────────────────
    uploaded_xlsx = st.file_uploader(
        "xlsx 파일 선택 (Normal + Easy + Difficult, 여러 스토리 동시 가능)",
        type=["xlsx"],
        accept_multiple_files=True,
        key="step2_files",
    )

    st.write("")

    col_run2, col_dl2 = st.columns([1, 1])

    with col_run2:
        run_btn2 = st.button("📄 TXT 생성", use_container_width=True, key="step2_run")

    if run_btn2:
        if not api_key:
            st.warning("⚠️ Gemini API 키를 먼저 입력해주세요.")
        elif not uploaded_xlsx:
            st.warning("⚠️ xlsx 파일을 업로드해주세요.")
        else:
            with st.spinner("Gemini로 감정 분석 중... 잠시 기다려주세요."):
                try:
                    from google import genai as _genai

                    client = _genai.Client(api_key=api_key)
                    MODEL  = "gemini-2.5-flash"

                    groups, unmatched = group_xlsx_files(uploaded_xlsx)

                    if unmatched:
                        st.warning(
                            f"파일명 형식을 인식할 수 없어 건너뜀: {', '.join(unmatched)}\n"
                            "파일명은 **ID_Title_Normal.xlsx** 형식이어야 합니다."
                        )
                    if not groups:
                        st.error("인식된 파일 그룹이 없습니다. 파일명을 확인해주세요.")
                        st.stop()

                    results: dict[str, bytes] = {}
                    prog = st.progress(0, text="처리 중...")

                    for i, (base, level_files) in enumerate(sorted(groups.items())):
                        prog.progress(
                            (i) / len(groups),
                            text=f"[{i+1}/{len(groups)}] {base} 처리 중...",
                        )

                        level_rows = {
                            lvl: read_xlsx_rows(data)
                            for lvl, data in level_files.items()
                            if lvl in LEVELS
                        }

                        normal_rows = level_rows.get("Normal", [])
                        scene_texts = extract_scene_texts(normal_rows)
                        scene_keys  = sorted(scene_texts.keys())

                        emotions = get_emotions(client, MODEL, scene_texts)
                        txt_str  = build_txt_content(level_rows, emotions, scene_keys)

                        parts    = base.split("_", 1)
                        sid      = parts[0]
                        title    = parts[1] if len(parts) > 1 else base
                        out_name = f"{sid}_{title}_Source.txt"
                        results[out_name] = txt_str.encode("utf-8")

                    prog.progress(1.0, text="완료!")

                    st.session_state.step2_result = results
                    st.success(f"✅ {len(results)}개 TXT 파일 생성 완료!")

                except Exception as e:
                    st.error(f"오류가 발생했습니다: {e}")
                    st.exception(e)
                    st.session_state.step2_result = None

    if st.session_state.step2_result:
        results = st.session_state.step2_result
        with col_dl2:
            if len(results) == 1:
                fname, content = next(iter(results.items()))
                st.download_button(
                    label=f"⬇️ {fname}",
                    data=content,
                    file_name=fname,
                    mime="text/plain",
                    use_container_width=True,
                    key="step2_dl",
                )
            else:
                zip_buf = io.BytesIO()
                with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                    for fname, content in results.items():
                        zf.writestr(fname, content)
                zip_buf.seek(0)
                st.download_button(
                    label="⬇️ ZIP 다운로드",
                    data=zip_buf.getvalue(),
                    file_name="txt_output.zip",
                    mime="application/zip",
                    use_container_width=True,
                    key="step2_dl",
                )


# ═════════════════════════════════════════════════════════
# STEP 3 — Chunk & Vocab
# ═════════════════════════════════════════════════════════
with tab3:
    st.subheader("Step 3. Chunk & Vocab txt (ICMS)")

    # ──────────────────────────────────────────────────────
    # Section 1: 파일 입력
    # ──────────────────────────────────────────────────────
    st.markdown("#### 📂 파일 입력")
    st.caption(
        "Step 2에서 생성된 **StoryID_StoryTitle_Source.txt** 파일을 업로드하세요. "
        "여러 스토리를 동시에 처리할 수 있습니다."
    )

    _secret_key3 = ""
    try:
        _secret_key3 = st.secrets.get("GEMINI_API_KEY", "")
    except Exception:
        pass

    if _secret_key3:
        api_key3 = _secret_key3
        st.success("✅ Gemini API 키가 설정되어 있습니다.", icon="🔑")
    else:
        api_key3 = st.text_input(
            "🔑 Gemini API Key",
            type="password",
            placeholder="AIza...",
            help="Chunk 구분 시에만 필요합니다. https://aistudio.google.com/app/apikey",
            key="step3_api_key",
        )

    uploaded_txts = st.file_uploader(
        "TXT 파일 선택 (여러 파일 동시 가능)",
        type=["txt"],
        accept_multiple_files=True,
        key="step3_files",
    )

    st.divider()

    # ──────────────────────────────────────────────────────
    # Section 2: Chunk
    # ──────────────────────────────────────────────────────
    st.markdown("#### ✂️ Chunk")
    st.markdown("""
<div style="background:#e8f4fd; border-left:4px solid #4F81BD;
            padding:10px 14px; border-radius:4px; font-size:0.82rem; line-height:1.7;">
📌 Thornbury (2019) 어휘 청크 기준 적용 — 관용표현·연어·구동사·담화 단위 등 식별 (Gemini API 사용)<br>
📌 청크 조건: 관습성·고정성·관용성·기능성 4가지 속성 중 2개 이상 충족 시 하나의 청크로 묶음<br>
📌 키 형식: <code>SC01_ST01_N_C</code> — 원래 키에 <code>_C</code>가 붙습니다.<br>
📌 구분자: <code>@@</code> (좌우 공백 없음)<br>
📌 출력 파일명: <code>StoryID_StoryTitle_C.txt</code>
</div>
""", unsafe_allow_html=True)

    st.write("")

    col_chunk_run, col_chunk_dl = st.columns([1, 1])
    with col_chunk_run:
        chunk_btn = st.button("✂️ Chunk 구분", use_container_width=True, key="step3_chunk_run")

    if chunk_btn:
        if not api_key3:
            st.warning("⚠️ Chunk 구분에는 Gemini API 키가 필요합니다.")
        elif not uploaded_txts:
            st.warning("⚠️ TXT 파일을 업로드해주세요.")
        else:
            with st.spinner("Gemini로 청크 분할 중... 잠시 기다려주세요."):
                try:
                    from google import genai as _genai
                    client3 = _genai.Client(api_key=api_key3)
                    MODEL3  = "gemini-2.5-flash"
                    chunk_results: dict[str, bytes] = {}
                    prog_c = st.progress(0, text="처리 중...")
                    for i, f in enumerate(uploaded_txts):
                        prog_c.progress(
                            i / len(uploaded_txts),
                            text=f"[{i+1}/{len(uploaded_txts)}] {f.name} 처리 중...",
                        )
                        f.seek(0)
                        raw   = f.read().decode("utf-8-sig")
                        pairs = parse_txt_for_chunking(raw)
                        if not pairs:
                            st.warning(f"⚠️ {f.name}: 처리할 문장이 없습니다.")
                            continue
                        chunked  = chunk_sentences_with_gemini(client3, MODEL3, pairs)
                        txt_out  = build_chunked_txt(pairs, chunked)
                        base     = re.sub(r"_Source\.txt$", "", f.name, flags=re.IGNORECASE)
                        base     = re.sub(r"\.txt$",        "", base,   flags=re.IGNORECASE)
                        chunk_results[base + "_C.txt"] = txt_out.encode("utf-8")
                    prog_c.progress(1.0, text="완료!")
                    st.session_state.step3_chunk_result = chunk_results
                    if chunk_results:
                        st.success(f"✅ {len(chunk_results)}개 파일 생성 완료!")
                    else:
                        st.error("처리된 파일이 없습니다. 업로드 파일을 확인해주세요.")
                except Exception as e:
                    st.error(f"오류가 발생했습니다: {e}")
                    st.exception(e)
                    st.session_state.step3_chunk_result = None

    if st.session_state.step3_chunk_result:
        cr = st.session_state.step3_chunk_result
        with col_chunk_dl:
            if len(cr) == 1:
                fn, fc = next(iter(cr.items()))
                st.download_button(
                    label=f"⬇️ {fn}",
                    data=fc,
                    file_name=fn,
                    mime="text/plain",
                    use_container_width=True,
                    key="step3_chunk_dl",
                )
            else:
                zb = io.BytesIO()
                with zipfile.ZipFile(zb, "w", zipfile.ZIP_DEFLATED) as zf:
                    for fn, fc in cr.items():
                        zf.writestr(fn, fc)
                zb.seek(0)
                st.download_button(
                    label="⬇️ ZIP 다운로드",
                    data=zb.getvalue(),
                    file_name="chunk_output.zip",
                    mime="application/zip",
                    use_container_width=True,
                    key="step3_chunk_dl",
                )

    st.divider()

    # ──────────────────────────────────────────────────────
    # Section 3: Vocab
    # ──────────────────────────────────────────────────────
    st.markdown("#### 📝 Vocab")
    st.markdown("""
<div style="background:#e8f4fd; border-left:4px solid #4F81BD;
            padding:10px 14px; border-radius:4px; font-size:0.82rem; line-height:1.7;">
📌 각 칸에 단어를 콤마(,)로 구분해 입력하면 해당 레벨 텍스트에서 단어를 찾아 <code>##단어##</code>로 표시합니다.<br>
📌 원형 입력 시 굴절형(복수·3인칭·과거·진행형 등)도 자동 인식됩니다. 불규칙 동사 포함.<br>
📌 Normal 단어 → <code>_N</code> 키 / Easy 단어 → <code>_E</code> 키 / Difficult 단어 → <code>_D</code> 키<br>
📌 출력 파일명: <code>StoryID_StoryTitle.txt</code>
</div>
""", unsafe_allow_html=True)

    vocab_n_raw = st.text_input(
        "Normal 단어", placeholder="planet, friend, watch ...", key="vocab_n"
    )
    vocab_e_raw = st.text_input(
        "Easy 단어", placeholder="planet, friend, watch ...", key="vocab_e"
    )
    vocab_d_raw = st.text_input(
        "Difficult 단어", placeholder="planet, friend, watch ...", key="vocab_d"
    )

    st.write("")

    col_vocab_run, col_vocab_dl = st.columns([1, 1])
    with col_vocab_run:
        vocab_btn = st.button("📝 Vocab 표기", use_container_width=True, key="step3_vocab_run")

    if vocab_btn:
        if not uploaded_txts:
            st.warning("⚠️ TXT 파일을 업로드해주세요.")
        else:
            try:
                vocab_n = [w.strip() for w in vocab_n_raw.split(",") if w.strip()]
                vocab_e = [w.strip() for w in vocab_e_raw.split(",") if w.strip()]
                vocab_d = [w.strip() for w in vocab_d_raw.split(",") if w.strip()]

                vocab_results: dict[str, bytes] = {}
                for f in uploaded_txts:
                    f.seek(0)
                    raw   = f.read().decode("utf-8-sig")
                    pairs = parse_txt_for_chunking(raw)
                    if not pairs:
                        st.warning(f"⚠️ {f.name}: 처리할 문장이 없습니다.")
                        continue
                    txt_out      = build_vocab_txt(pairs, vocab_n, vocab_e, vocab_d)
                    emotion_lines = parse_emotion_lines(raw)
                    if emotion_lines:
                        txt_out = txt_out + "\n\n" + "\n".join(emotion_lines)
                    base     = re.sub(r"_Source\.txt$", "", f.name, flags=re.IGNORECASE)
                    base     = re.sub(r"\.txt$",        "", base,   flags=re.IGNORECASE)
                    vocab_results[base + ".txt"] = txt_out.encode("utf-8")

                st.session_state.step3_vocab_result = vocab_results
                if vocab_results:
                    st.success(f"✅ {len(vocab_results)}개 파일 생성 완료!")
                else:
                    st.error("처리된 파일이 없습니다.")
            except Exception as e:
                st.error(f"오류가 발생했습니다: {e}")
                st.exception(e)
                st.session_state.step3_vocab_result = None

    if st.session_state.step3_vocab_result:
        vr = st.session_state.step3_vocab_result
        with col_vocab_dl:
            if len(vr) == 1:
                fn, fc = next(iter(vr.items()))
                st.download_button(
                    label=f"⬇️ {fn}",
                    data=fc,
                    file_name=fn,
                    mime="text/plain",
                    use_container_width=True,
                    key="step3_vocab_dl",
                )
            else:
                zb = io.BytesIO()
                with zipfile.ZipFile(zb, "w", zipfile.ZIP_DEFLATED) as zf:
                    for fn, fc in vr.items():
                        zf.writestr(fn, fc)
                zb.seek(0)
                st.download_button(
                    label="⬇️ ZIP 다운로드",
                    data=zb.getvalue(),
                    file_name="vocab_output.zip",
                    mime="application/zip",
                    use_container_width=True,
                    key="step3_vocab_dl",
                )
